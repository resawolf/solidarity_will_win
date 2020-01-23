# Left to do: SChriftgröße gui, mails leerzeichen, schreiben in excel-sheet Nötig: install xlrd, pandas, sys, tkinter

#Pad herunterladen ( https://ethercalc.org/zw1njrpgh5fw)
import urllib.request
url = "https://ethercalc.org/zw1njrpgh5fw.xlsx"
urllib.request.urlretrieve(url, 'C:/Users/teres/OneDrive/Desktop/Dokumente/Python/Stuttii2.xlsx')

# Excelsheet importieren
import xlrd
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook
file_location = "C:/Users/teres/OneDrive/Desktop/Dokumente/Python/Stuttii2.xlsx"
workbook = xlrd.open_workbook(file_location)
sheet = workbook.sheet_by_index(0)
namensliste = sheet.col_values(0, start_rowx=31, end_rowx=36)
laenge = len(namensliste)
emailliste = sheet.col_values(1, start_rowx=31, end_rowx=36)

# Block für die Vorbereitung der Erinnerungsmails
import sys, smtplib, os
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.utils import formatdate
server = smtplib.SMTP("smtp.de.posteo.de", 587)
senderEmail = "teresa.erbach@posteo.de"

# Tkinter importieren
from tkinter import *
import tkinter.messagebox

# Hauptfenster erstellen
root = Tk()
Monat = StringVar()

# Eigenschaften fuer das Fenster setzen
root.title("Stuttobotter")
root.minsize(width = 1350, height = 740)

#Hier noch Leerzeile einfügen # Schrift größer # Farben
Begruessung = Label(root,text="Hallo! Für welchen Monat möchtest du die Abrechnung machen?")
Begruessung.pack(pady = 10)
   
Monat = StringVar()
Monatseingabe = Entry(root, textvariable = Monat)
Monatseingabe.pack(pady = 10)

liste = ["Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember", "13.Monat"]

#Ausgabensummen eintragen
for i in range(1,13):
    summe = sheet.cell_value(i,1) + sheet.cell_value(i,2) + sheet.cell_value(i,3) + sheet.cell_value(i,4) + sheet.cell_value(i,5)
    wb2 = load_workbook('C:/Users/teres/OneDrive/Desktop/Dokumente/Python/Stuttii2.xlsx')
    ws4 = wb2["Sheet1"]
    d = ws4.cell(row= i+1, column= 7, value=summe)
    wb2.save('C:/Users/teres/OneDrive/Desktop/Dokumente/Python/Stuttii2.xlsx')

#Gebotesummen eintragen
for i in range(1,13):
    summe = sheet.cell_value(i,9) + sheet.cell_value(i,10) + sheet.cell_value(i,11) + sheet.cell_value(i,12) + sheet.cell_value(i,13)
    wb2 = load_workbook('C:/Users/teres/OneDrive/Desktop/Dokumente/Python/Stuttii2.xlsx')
    ws4 = wb2["Sheet1"]
    d = ws4.cell(row= i+1, column= 15, value=summe)
    wb2.save('C:/Users/teres/OneDrive/Desktop/Dokumente/Python/Stuttii2.xlsx')

#Nonfoodsumme eintragen
for i in range(15,27):
    summe = sheet.cell_value(i,1) + sheet.cell_value(i,2) + sheet.cell_value(i,3) + sheet.cell_value(i,4) + sheet.cell_value(i,5)
    wb2 = load_workbook('C:/Users/teres/OneDrive/Desktop/Dokumente/Python/Stuttii2.xlsx')
    ws4 = wb2["Sheet1"]
    d = ws4.cell(row= i+1, column= 7, value=summe)
    wb2.save('C:/Users/teres/OneDrive/Desktop/Dokumente/Python/Stuttii2.xlsx')

#Lesemodus
file_location = "C:/Users/teres/OneDrive/Desktop/Dokumente/Python/Stuttii2.xlsx"
workbook = xlrd.open_workbook(file_location)
sheet = workbook.sheet_by_index(0)

def Mo_auslesen():
    while (Monat.get() not in liste):
        Start = Label(root, text = "Nee... Monatsangabe großschreiben und keine Jahreszahl!" + ":\n" )
        Start.pack()
        break

    if (Monat.get() in liste):
        Start = Label(root, text = "\nAbrechnung für " + Monat.get() + ":\n" )
        Start.pack()
        indices = [i for i, x in enumerate(liste) if x == Monat.get()]  # Der wievielte Monat in der Liste ist das?
        a = indices[0] + 1  # Die Indices werden als Liste ausgegeben; a ist das erste Element davon

        # Ausgabensumme auslesen
        for col_index, cell in enumerate(sheet.row_values(0)):
            if cell == "Gesamtausgaben":
                g = col_index
        summe = sheet.cell_value(a,g)

        # Gebotesumme auslesen 
        for col_index, cell in enumerate(sheet.row_values(0)):
            if cell == "Gebotesumme":
                l = col_index    
        gebotesumme = sheet.cell_value(a, l)
    
        differenz = summe - gebotesumme
        

        # Nonfoodsumme auslesen
        for col_index, cell in enumerate(sheet.row_values(14)):
            if cell == "Summe_Nonfood":
                l = col_index
        b = a + 14  
        nonfoodsumme = sheet.cell_value(b, l)
        nonf_ind = nonfoodsumme / laenge

        # Mail vorbereiten       
        server.starttls()
        server.login("teresa.erbach@posteo.de", "Somnambul6.")

    def Check(i,b):
        k = i + 1
        if k in range(0, laenge):
            for col_index, cell in enumerate(sheet.row_values(0)):
                if cell == str(namensliste[k]):
                    j = col_index
                    if b ==1:
                        break


                def weiter():
                    Fenster_i.destroy()
                    Check(k,b)

                def Mail_verschicken():
                    server = smtplib.SMTP("smtp.de.posteo.de", 587)
                    senderEmail = "teresa.erbach@posteo.de"
                    server.starttls()
                    server.login("teresa.erbach@posteo.de", "Somnambul6.")
                    msg = MIMEMultipart()
                    msg["From"] = senderEmail
                    msg["Subject"] = "Das ist"       
                    empfangsEmail = str(emailliste[k])
                    msg["To"] = empfangsEmail

                    if b == 1:
                        emailText = "eine Essensausgabeneintragungserinnerungsmail fuer den Monat " + Monat.get()
                    if b == 0:
                        emailText = "eine Bieteverfahrenerinnerungsmail fuer den Monat " + Monat.get()

                    msg.attach(MIMEText(emailText, "html"))     
                    text = msg.as_string()
          
                    try:
                        server.sendmail(senderEmail, empfangsEmail, text)
                        tkinter.messagebox.showinfo("\nMail schon verschickt!\n")
                        Fenster_i.destroy()                        
                        Check(k, b)

                    except:
                        tkinter.messagebox.showinfo("Hat leider nicht funktioniert. Please try again later! :) ")
                        Fenster_i.destroy()
                        Check(k, b)

                    del msg
                    server.quit()


            y = (sheet.cell_type(a,j))
            if (y == 0) or (sheet.cell_value(a,j)==0):
                if b == 1:
                    v = "keine Ausgaben eingetragen."
                    h = 200
                if b == 0:
                    v = "kein Gebot eingetragen."
                    h = 300

                Fenster_i = Tk()
                Fenster_i.title(str(namensliste[k]))
                Fenster_i.geometry('%dx%d+%d+%d' % (300, 120, 520, h))

                Mailabfrage = Label(Fenster_i, text = (str(namensliste[k]) + " hat noch " + v + "\n Möchtest du " + str(namensliste[k]) + " jetzt eine Erinnerungsmail schicken?"))
                Mailabfrage.pack()

                Bestätigung = Button(Fenster_i, text = "ja", command = Mail_verschicken)
                Bestätigung.pack(pady = 10)

                Nein = Button(Fenster_i, text = "nein", command = weiter)
                Nein.pack(pady = 10)

            else:
                Check(k,b)

        else:                
            Test = False
            for i in range(0, laenge):
                for col_index, cell in enumerate(sheet.row_values(0)):
                    if cell == str(namensliste[i]):
                        c = col_index
                        if b == 1:
                            break
                s = sheet.cell_type(a,c)
                if (s == 0) or (sheet.cell_value(a,c)==0):
                    Test = True                                                   
                
            if Test == True: # d. h. es fehlen noch /Gebote
                Ende1 = Label (root, text = "\n \n \n \n \n \n \n \n Da musst du jetzt wohl auf die Schluffis warten... Geh nach draußen spielen!\n")
                Ende1.pack()
                def logout(): 
                    quit()
                Logout = Button(root, text = "logout", command = logout)
                Logout.pack(pady = 10)

            else:# d. h. alle haben ihre Ausgaben/Gebote eingetragen
                if b == 1: # Ausgaben
                    i = 0
                    Gesamtausgaben = Label(root, text = ("\nInsgesamt wurden im " + Monat.get() + " " + str(summe) + " Euro ausgegeben. \n \n Überprüfung, ob schon alle geboten haben:"))
                    Gesamtausgaben.pack()
                    Check(i,0)

                else: # Gebote 
                    if differenz <= 0:
                        dif = str(-differenz) + " Euronen zu viel geboten.\n\n "
                    else:
                        dif= str( differenz) + " Euronen zu wenig geboten.\n\n "

                    Geboteanzeige = Label(root, text = "\n \n Haben alle gemacht - insgesamt wurden " + dif)
                    Geboteanzeige.pack()

                    def logout():
                            exit()

                    def Ende():
                        Ende = Label (root, text = "\n \n \n \n \n \n  Hab einen schönen Tag! \n \n Viele Küsse, dein Stuttoboter")
                        Ende.pack()
                        Logout = Button(root, text = "logout", command = logout)
                        Logout.pack(pady = 10)

                    def Infomails(e):

                        Info = Tk()
                        Info.title("Infomails")
                        Info.geometry('%dx%d+%d+%d' % (500, 120, 450, 400))
                        Abfrage_Infomails = Label(Info, text = ("Sollen Infomails mit den zu zahlenden Beträgen an alle verschickt werden?"))
                        Abfrage_Infomails.pack()
                        
                        def Ende1():
                            Info.destroy()
                            Keine_mails = Label(root, text = "\n \n Keine Mails verschickt.")
                            Keine_mails.pack()
                            Ende()                          
                        
                        
                        def Ende2():
                            Info.destroy()                                  
                            Test = True                       
                        
                            for i in range(0, laenge):
                                for col_index, cell in enumerate(sheet.row_values(0)):
                                    if cell == str(namensliste[i]):
                                        k = col_index
                                        break                                            

                                msg = MIMEMultipart()

            
                                if e == 1:  # hier keine Anpassung der Gebote                                
                                    endbetrag = sheet.cell_value(a, k+8) + nonf_ind - sheet.cell_value(a, k) - sheet.cell_value(a + 14, k)
                                    endbetrag1 = round(endbetrag, 2)
                                    emailText = None 
                                    if endbetrag1 > 0:
                                        emailText = "Liebe/r " + str(namensliste[i]) + ", <br><br>alle haben für den " + Monat.get() + " genug geboten. Es wäre großartig, wenn du möglichst bald " + str(endbetrag1) + " Euro aufs WG-Konto überweist. (Setzt sich zusammen aus deinem Gebot von " + str(sheet.cell_value(a, k+8)) + " Euro plus " + str(nonf_ind) + " Euro Nonfood-Anteil abzüglich deiner Ausgaben von" + str(sheet.cell_value(a, k)) + " Euro für Essen und " + str(sheet.cell_value(a+14, k)) + " Euro für Sonstiges.)<br><br> Allerliebst, der Stuttoboter"
                                    if endbetrag1 < 0:
                                        emailText= "Liebe/r " + str(namensliste[i]) + ", <br><br>alle haben für den " + Monat.get() + " genug geboten. Deine Ausgaben waren höher als dein Gebot und daher kriegst du " + str(-endbetrag1) + " Euro zurücküberwiesen. (Setzt sich zusammen aus deinem Gebot von " + str(sheet.cell_value(a, k+8)) + " Euro plus " + str(nonf_ind) + " Euro Nonfood-Anteil abzüglich Ausgaben von" + str(sheet.cell_value(a, k)) + " Euro für Essen und " + str(sheet.cell_value(a+14, k)) + " Euro für Sonstiges.)<br><br> Allerliebst, der Stuttoboter"
                                    if endbetrag1 == 0:
                                        emailText = "Liebe/r " + str(namensliste[i]) + ", <br><br>alle haben für den " + Monat.get() + " genug geboten. Deine Ausgaben waren genauso hoch wie dein Gebot und du musst daher nichts überweisen. (Setzt sich zusammen aus deinem Gebot von " + str(sheet.cell_value(a, k+8)) + " Euro plus " + str(nonf_ind) + " Euro Nonfood-Anteil abzüglich deiner Ausgaben von" + str(sheet.cell_value(a, k)) + " Euro für Essen und " + str(sheet.cell_value(a+14, k)) + " Euro für Sonstiges.)<br><br> Allerliebst, der Stuttoboter"

                                    
                                else:# Anpassung der Gebote)
                                    gebot_angepasst = sheet.cell_value(a, k+8) + ((differenz) * (sheet.cell_value(a, k+8) / gebotesumme))
                                    endbetrag = gebot_angepasst + nonf_ind - sheet.cell_value(a, k) - sheet.cell_value(a+14, k)
                                    endbetrag1 = round(endbetrag, 2)
                                    emailText = None
                                    if differenz <=0:
                                        aa = str(-differenz)
                                        ab = "zu viel"
                                        ac = "runtergesetzt."

                                    if differenz > 0:
                                        aa = str(differenz)
                                        ab = "zu wenig"
                                        ac = "erhöht."

                                    if endbetrag1 > 0:
                                        emailText = "Liebe/r " + str(namensliste[i]) + ", <br><br>insgesamt wurden bei der Bieterrunde für den " + Monat.get() + aa + " Euro " + ab + " geboten. Dein Gebot von " + str(sheet.cell_value(a, k+8)) + " wurde deshalb auf " + str(gebot_angepasst) + " Euro " + ac + " <br> Es wäre großartig, wenn du möglichst bald " + str(endbetrag1) + " Euro aufs WG-Konto überweist. (Setzt sich zusammen aus dem angepassten Gebot von " + str(gebot_angepasst) + " Euro plus " + str(nonf_ind) + " Euro Nonfood-Anteil abzüglich deiner Ausgaben von " + str(sheet.cell_value(a, k)) + " Euro für Essen und " + str(sheet.cell_value(a+14, k)) + " Euro für Sonstiges.)<br><br> Allerliebst, der Stuttoboter"
                                    if endbetrag1 == 0: 
                                        emailText = "Liebe/r " + str(namensliste[i]) + ", <br><br>insgesamt wurden bei der Bieterrunde für den " + Monat.get() + aa + " Euro " + ab + " geboten. Dein Gebot von " + str(sheet.cell_value(a, k+8)) + " wurde deshalb auf " + str(gebot_angepasst) + " Euro " + ac + "<br> Zuzüglich des Nonfood-Anteils von " + str(nonf_ind) + " Euro und abzüglich deiner Ausgaben von " + str(sheet.cell_value(a, k)) + " Euro für Essen und " + str(sheet.cell_value(a+14, k)) + " Euro für Sonstiges kommt genau 0 raus - du brauchst also nichts überweisen. <br><br> Allerliebst, der Stuttobotter"
                                    if endbetrag1 <0:
                                        emailText = "Liebe/r " + str(namensliste[i]) + ", <br><br>insgesamt wurden bei der Bieterrunde für den " + Monat.get() + aa + " Euro " + ab + " geboten. Dein Gebot von " + str(sheet.cell_value(a, k+8)) + " wurde deshalb auf " + str(gebot_angepasst) + " Euro " + ac + "Da deine Ausgaben höher waren, bekommst du vom WG-Konto " + str(-endbetrag1) + " Euro zurücküberwiesen. (Setzt sich zusammen aus dem angepassten Gebot von " + str(gebot_angepasst) + " Euro plus " + str(nonf_ind) + " Euro Nonfood-Anteil abzüglich deiner Ausgaben von " + str(sheet.cell_value(a, k)) + " Euro für Essen und " + str(sheet.cell_value(a+14, k)) + " Euro für Sonstiges.)<br><br> Allerliebst, der Stuttobotter"
                                    
                                                                   
                           
                                try:  
                                    server = smtplib.SMTP("smtp.de.posteo.de", 587)
                                    senderEmail = "teresa.erbach@posteo.de"
                                    server.starttls()
                                    server.login("teresa.erbach@posteo.de", "Somnambul6.")
                                    msg = MIMEMultipart()
                                    msg["From"] = senderEmail
                                    msg["Subject"] = "Überweisung Essensgeld"        
                                    empfangsEmail = str(emailliste[i])
                                    msg["To"] = empfangsEmail
                                    msg.attach(MIMEText(emailText, "html"))
                                    text = msg.as_string()
                                    server.sendmail(senderEmail, empfangsEmail, text)
                                    server.quit()                                                                            
                                        
                                except:
                                    tkinter.messagebox.showinfo("FEHLER IM SYSTEM! \n Mail an " + str(namensliste[i]) + " konnte nicht verschickt werden!")
                                    Test = False  
                                    
                                del msg
                                if endbetrag1 < 0:
                                   
                                    try:
                                        server = smtplib.SMTP("smtp.de.posteo.de", 587)
                                        senderEmail = "teresa.erbach@posteo.de"
                                        server.starttls()
                                        server.login("teresa.erbach@posteo.de", "Somnambul6.")
                                        msg = MIMEMultipart()
                                        msg["From"] = senderEmail
                                        msg["Subject"] = "Überweisung Essensgeld"          
                                        empfangsEmail = senderEmail
                                        msg["To"] = empfangsEmail
                                        emailText2 = "Bitte überweise für den Monat " + Monat.get() + str(-endbetrag1) + " Euro vom WG-Konto an " + str(namensliste[i])+ ". <br><br> Allerliebst, dein STuttobotter"
                                        msg.attach(MIMEText(emailText2, "html")) 
                                        text2 = msg.as_string()
                                        server.sendmail(senderEmail, empfangsEmail, text2)
                                        server.quit()
                                    except:
                                        tkinter.messagebox.showinfo("FEHLER IM SYSTEM! \n Mail an teresa.erbach@posteo.de konnte nicht verschickt werden!")
                                        Test = False    
                                    del msg                    
                                                                    
                            if Test == True:
                                Bestätigung = Label(root, text = "Mails wurden an alle verschickt!")
                                Bestätigung.pack()
                            else:
                                Bestätigung = Label(root, text = "Mails an alle anderen verschickt!")
                                Bestätigung.pack()

                            Ende()

                        
                        Info_ja = Button(Info, text = "ja", command = Ende2)
                        Info_ja.pack(pady = 10)
                        Info_nein = Button(Info, text = "nein", command = Ende1)
                        Info_nein.pack(pady = 10)
                       

                    def Gebotsanpassung():
                        Gebotsangleich.destroy()
                        
                        for i in range(0, laenge):
                            for col_index, cell in enumerate(sheet.row_values(0)):
                                if cell == str(namensliste[i]):
                                    k = col_index 
                                    break
                            gebot_angepasst = sheet.cell_value(a, k+8) + ((differenz) * (sheet.cell_value(a, k+8) / gebotesumme))
                            endbetrag = float(gebot_angepasst) + float(nonf_ind) - sheet.cell_value(a, k) - sheet.cell_value(a+14, k)
                            endbetrag1 = round(endbetrag, 2)  

                            wb2 = load_workbook('C:/Users/teres/OneDrive/Desktop/Dokumente/Python/Stuttii2.xlsx')
                            ws4 = wb2["Sheet1"]
                            d = ws4.cell(row= a+15, column= k+9, value=endbetrag1)
                            wb2.save('C:/Users/teres/OneDrive/Desktop/Dokumente/Python/Stuttii2.xlsx')

                            #Eintrag in Excel-Tabelle
                                                              

                        Gebotsanpassung = Label(root, text = "Gebote wurden angepasst und in die Exceltabelle eingetragen.")
                        Gebotsanpassung.pack()
                        Infomails(0)

                    def Keine_Gebotsanpassung():
                        Gebotsangleich.destroy()

                        for i in range(0, laenge):
                            for col_index, cell in enumerate(sheet.row_values(0)):
                                if cell == str(namensliste[i]):
                                    k = col_index
                                    break
                           
                            endbetrag = sheet.cell_value(a, k+8) + int(nonf_ind) - sheet.cell_value(a, k) - sheet.cell_value(a + 14, k)
                            endbetrag1 = round(endbetrag, 2)
                                                                                                                                       
                                        #Eintrag in Excel-Tabell       
                            wb2 = load_workbook('C:/Users/teres/OneDrive/Desktop/Dokumente/Python/Stuttii2.xlsx')
                            ws4 = wb2["Sheet1"]
                            d = ws4.cell(row= a+15, column= k+9, value=endbetrag1)
                            wb2.save('C:/Users/teres/OneDrive/Desktop/Dokumente/Python/Stuttii2.xlsx')
                        
                        Gebotsanpassung = Label(root, text = "Gebote werden nicht angepasst.")
                        Gebotsanpassung.pack()
                        Infomails(1)

                    Gebotsangleich = Tk()
                    Gebotsangleich.title("Anpassung der Gebote")
                    Gebotsangleich.geometry('%dx%d+%d+%d' % (300, 120, 520, 350))
                    Abfrage_Gebotsangleich = Label(Gebotsangleich, text = ("Sollen die Gebote prozentual angepasst werden?"))
                    Abfrage_Gebotsangleich.pack()
                    Bestätigung = Button(Gebotsangleich, text = "ja", command = Gebotsanpassung )
                    Bestätigung.pack(pady = 10)
                    Nein = Button(Gebotsangleich, text = "nein", command = Keine_Gebotsanpassung)
                    Nein.pack(pady = 10)                            
                                             

    Check(-1,1)


Bestätigung = Button(root, text = "weiter", command = Mo_auslesen)
Bestätigung.pack(pady = 10)

root.mainloop()
